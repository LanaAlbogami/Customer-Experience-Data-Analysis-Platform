import pandas as pd
import io
import os
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook

# =====================================================================
# 1. TEXT PROCESSING & FORMATTING UTILITIES
# =====================================================================

def process_arabic_text(text):
    """
    Step 1: Convert the incoming input into a standard string.
    Step 2: Reshape the Arabic letters using 'arabic_reshaper' so they connect properly.
    Step 3: Apply the BiDi algorithm via 'get_display' to reverse the text direction for correct PDF rendering.
    Step 4: Return the final properly formatted Arabic text string.
    """
    reshaped_text = arabic_reshaper.reshape(str(text))
    bidi_text = get_display(reshaped_text)
    return bidi_text

def format_val(val, is_csat=False):
    """
    Step 1: Try to convert the given value into a floating-point number.
    Step 2: If successful, format it cleanly to 2 decimal places.
    Step 3: Check if the value belongs to CSAT; if true, append '%' sign.
    Step 4: If conversion fails (e.g., None or text), return the original value as a string.
    """
    try:
        num = float(val)
        formatted = f"{num:.2f}"
        return f"{formatted}%" if is_csat else formatted
    except:
        return str(val)


# =====================================================================
# 2. DATA MATRIX TRANSFORMATION
# =====================================================================

def build_service_matrix(rows):
    """
    Step 1: Initialize an empty dictionary to aggregate service metrics uniquely.
    Step 2: Iterate through every raw database record row provided.
    Step 3: Extract department name, service name, indicator name, previous value, and current value.
    Step 4: Create a unique composite key combining department and service name.
    Step 5: If the key doesn't exist in the dictionary, initialize default baseline values for CSAT, CES, and NPS.
    Step 6: Clean and convert the indicator name to uppercase to safely check its category.
    Step 7: Check if the indicator is CSAT, CES, or NPS, then update its corresponding previous, target, and current values.
    Step 8: Convert the aggregated dictionary values into a clean list and return it.
    """
    services_dict = {}
    
    # Loop through each row in the dataset
    for row in rows:
        dept = row['DepartmentName']
        s = row['ServiceName']
        ind = row['IndicatorName']
        p = row['PrevValue']
        c = row['CurrentValue']
        
        # Create a unique key for grouping by department and service
        key = (dept, s)
        
        # Initialize default structure if the service is encountered for the first time
        if key not in services_dict:
            services_dict[key] = {
                "Department": dept,
                "Service": s,
                "Prvious CSAT": "-", "Target CSAT": format_val(85, is_csat=True), "Current CSAT": "-",
                "Prvious CES": "-", "Target CES": format_val(76, is_csat=False), "Current CES": "-",
                "Prvious NPS": "-", "Target NPS": format_val(69, is_csat=False), "Current NPS": "-"
            }
        
        # Clean the indicator string for robust keyword matching
        ind_clean = str(ind).strip().upper()
        
        # Map values into their correct metric columns based on indicator type
        if "CSAT" in ind_clean:
            services_dict[key]["Prvious CSAT"] = format_val(p, is_csat=True)
            services_dict[key]["Target CSAT"] = format_val(85, is_csat=True)
            services_dict[key]["Current CSAT"] = format_val(c, is_csat=True)
        elif "CES" in ind_clean:
            services_dict[key]["Prvious CES"] = format_val(p, is_csat=False)
            services_dict[key]["Target CES"] = format_val(76, is_csat=False)
            services_dict[key]["Current CES"] = format_val(c, is_csat=False)
        elif "NPS" in ind_clean:
            services_dict[key]["Prvious NPS"] = format_val(p, is_csat=False)
            services_dict[key]["Target NPS"] = format_val(69, is_csat=False)
            services_dict[key]["Current NPS"] = format_val(c, is_csat=False)
            
    # Return the final structured rows as a list
    return list(services_dict.values())


# =====================================================================
# 3. EXCEL REPORT GENERATION
# =====================================================================

def generate_excel(period, periods_data):
    """
    Step 1: Create an in-memory byte buffer to hold the generated Excel file.
    Step 2: Initialize pandas ExcelWriter using openpyxl engine.
    Step 3: Check if multiple periods are selected; if so, calculate and build an annual summary sheet.
    Step 4: Loop through each period data dictionary and create individual Excel sheets.
    Step 5: Load the saved workbook via openpyxl to apply professional header design and styling.
    Step 6: Save the final styled workbook into the buffer and return its binary content, filename, and mime type.
    """
    buffer = io.BytesIO()
    
    # Open excel writer context
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        selected_periods_list = list(periods_data.items())
        
        # --- Section A: Annual Summary Sheet Construction ---
        if len(selected_periods_list) > 1:
            annual_rows = []
            
            # Gather all unique services across all periods
            all_services = {}
            for p_name, rows in periods_data.items():
                for row in rows:
                    key = (row['DepartmentName'], row['ServiceName'])
                    if key not in all_services:
                        all_services[key] = {"dept": row['DepartmentName'], "service": row['ServiceName']}

            # Extract distinct years present in the dataset rows or period titles
            years_in_data = set()
            for p_name, rows in periods_data.items():
                for row in rows:
                    if 'Year' in row and row['Year']:
                        years_in_data.add(str(row['Year']))
            
            # Fallback parsing of year from period title string if row year is missing
            if not years_in_data:
                for p_name, _ in periods_data.items():
                    parts = p_name.split(" - ")
                    if parts:
                        years_in_data.add(parts[0].strip())

            # Loop through each distinct year to aggregate yearly averages
            for year_str in sorted(years_in_data):
                try:
                    prev_year_str = str(int(year_str) - 1)
                except:
                    prev_year_str = ""

                for (dept, s), info in all_services.items():
                    curr_csat_vals, curr_ces_vals, curr_nps_vals = [], [], []
                    prev_csat_vals, prev_ces_vals, prev_nps_vals = [], [], []

                    for p_name, rows in periods_data.items():
                        for row in rows:
                            if row['DepartmentName'] == dept and row['ServiceName'] == s:
                                row_year = str(row.get('Year', ''))
                                if not row_year:
                                    if p_name.startswith(year_str):
                                        row_year = year_str
                                    elif prev_year_str and p_name.startswith(prev_year_str):
                                        row_year = prev_year_str

                                ind_clean = str(row['IndicatorName']).strip().upper()
                                try:
                                    c_val = float(row['CurrentValue'])
                                except:
                                    continue

                                # Distribute values into current or previous year lists
                                if row_year == year_str:
                                    if "CSAT" in ind_clean: curr_csat_vals.append(c_val)
                                    elif "CES" in ind_clean: curr_ces_vals.append(c_val)
                                    elif "NPS" in ind_clean: curr_nps_vals.append(c_val)
                                elif row_year == prev_year_str:
                                    if "CSAT" in ind_clean: prev_csat_vals.append(c_val)
                                    elif "CES" in ind_clean: prev_ces_vals.append(c_val)
                                    elif "NPS" in ind_clean: prev_nps_vals.append(c_val)

                    # Helper function to compute safe mathematical average
                    def avg(lst, is_c=False):
                        if not lst:
                            return "-"
                        return format_val(sum(lst) / len(lst), is_csat=is_c)

                    # Append compiled annual row if valid data exists
                    if curr_csat_vals or curr_ces_vals or curr_nps_vals:
                        annual_rows.append({
                            "Year": year_str,
                            "Department": dept,
                            "Service": s,
                            "Prvious CSAT": avg(prev_csat_vals, True), 
                            "Target CSAT": format_val(85, is_csat=True), 
                            "Current CSAT": avg(curr_csat_vals, True),
                            "Prvious CES": avg(prev_ces_vals, False), 
                            "Target CES": format_val(76, is_csat=False), 
                            "Current CES": avg(curr_ces_vals, False),
                            "Prvious NPS": avg(prev_ces_vals, False), 
                            "Target NPS": format_val(69, is_csat=False), 
                            "Current NPS": avg(curr_nps_vals, False)
                        })

            # Write annual summary dataframe to excel sheet if rows are available
            if annual_rows:
                df_annual = pd.DataFrame(annual_rows)
                df_annual.to_excel(writer, index=False, sheet_name='ملخص السنة')

        # --- Section B: Individual Period Sheets Construction ---
        for p_name, rows in periods_data.items():
            matrix = build_service_matrix(rows)
            df_period = pd.DataFrame(matrix)
            safe_sheet_name = p_name.replace(" & ", "_")
            df_period.to_excel(writer, index=False, sheet_name=safe_sheet_name)

    # --- Section C: Workbook Styling (Headers Fill, Fonts, Alignment) ---
    buffer.seek(0)
    wb = load_workbook(buffer)
    
    header_fill = PatternFill(start_color="7A7F94", end_color="7A7F94", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Apply styles across all worksheets in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    
    return final_buffer.getvalue(), f"Report_{period.replace(' ', '_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# =====================================================================
# 4. PDF REPORT GENERATION
# =====================================================================

def generate_pdf(period, periods_data):
    """
    Step 1: Initialize an in-memory byte buffer for PDF document output.
    Step 2: Set up ReportLab SimpleDocTemplate with precise page dimensions and margins.
    Step 3: Register custom TrueType Arabic font (Arial) to support Arabic characters.
    Step 4: Define canvas header-footer callback to draw logos and running titles dynamically.
    Step 5: Process annual summary data and compile table elements if multiple periods are selected.
    Step 6: Loop through each period dataset, transform matrix, and append styled ReportLab tables.
    Step 7: Build the entire PDF document flow and return its byte content, filename, and mime type.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(842, 595), rightMargin=30, leftMargin=30, topMargin=85, bottomMargin=30)
    elements = []
    
    # Path configuration for system Arabic font registration
    font_path = "C:\\Windows\\Fonts\\arial.ttf"
    font_name = "Helvetica" 
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
        font_name = 'ArabicFont'
        
    styles = getSampleStyleSheet()
    
    # Callback function to render static headers and footers on pages
    def add_header_footer(canvas_obj, document):
        canvas_obj.saveState()
        logo_path = "logo.png"
        if os.path.exists(logo_path):
            canvas_obj.drawImage(logo_path, 40, 525, width=130, height=50, preserveAspectRatio=True, mask='auto')
        
        canvas_obj.setFont(font_name, 13)
        canvas_obj.drawRightString(800, 550, process_arabic_text(f"منصة تجربة العميل - التقرير الشامل للفترات: {period}"))
        
        canvas_obj.setLineWidth(1)
        canvas_obj.setStrokeColor(colors.HexColor('#1e253c'))
        canvas_obj.line(40, 515, 800, 515)
        canvas_obj.restoreState()

    elements.append(Spacer(1, 10))

    # Styling for section titles
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=11,
        alignment=2,
        textColor=colors.HexColor('#7b5ab8'),
        spaceAfter=8,
        spaceBefore=15,
        keepWithNext=True
    )

    # Define table column header labels and layout widths
    headers = [
        process_arabic_text("Current NPS"), process_arabic_text("Target NPS"), process_arabic_text("Prvious NPS"),
        process_arabic_text("Current CES"), process_arabic_text("Target CES"), process_arabic_text("Prvious CES"),
        process_arabic_text("Current CSAT"), process_arabic_text("Target CSAT"), process_arabic_text("Prvious CSAT"),
        process_arabic_text("Service"), process_arabic_text("Department")
    ]
    col_widths = [65, 65, 65, 65, 65, 65, 65, 65, 65, 90, 90]

    selected_periods_list = list(periods_data.keys())

    # --- Section A: Build PDF Annual Summary Table (If multiple periods chosen) ---
    if len(selected_periods_list) > 1:
        annual_table_data = [headers]
        
        all_services = {}
        for p_name, rows in periods_data.items():
            for row in rows:
                key = (row['DepartmentName'], row['ServiceName'])
                if key not in all_services:
                    all_services[key] = True

        years_in_data = set()
        for p_name, rows in periods_data.items():
            for row in rows:
                if 'Year' in row and row['Year']:
                    years_in_data.add(str(row['Year']))
        
        if not years_in_data:
            for p_name, _ in periods_data.items():
                parts = p_name.split(" - ")
                if parts:
                    years_in_data.add(parts[0].strip())

        for year_str in sorted(years_in_data):
            try:
                prev_year_str = str(int(year_str) - 1)
            except:
                prev_year_str = ""

            for (dept, s) in all_services.keys():
                curr_csat_vals, curr_ces_vals, curr_nps_vals = [], [], []
                prev_csat_vals, prev_ces_vals, prev_nps_vals = [], [], []

                for p_name, rows in periods_data.items():
                    for row in rows:
                        if row['DepartmentName'] == dept and row['ServiceName'] == s:
                            row_year = str(row.get('Year', ''))
                            if not row_year:
                                if p_name.startswith(year_str):
                                    row_year = year_str
                                elif prev_year_str and p_name.startswith(prev_year_str):
                                    row_year = prev_year_str

                            ind_clean = str(row['IndicatorName']).strip().upper()
                            try:
                                c_val = float(row['CurrentValue'])
                            except:
                                continue

                            if row_year == year_str:
                                if "CSAT" in ind_clean: curr_csat_vals.append(c_val)
                                elif "CES" in ind_clean: curr_ces_vals.append(c_val)
                                elif "NPS" in ind_clean: curr_nps_vals.append(c_val)
                            elif row_year == prev_year_str:
                                if "CSAT" in ind_clean: prev_csat_vals.append(c_val)
                                elif "CES" in ind_clean: prev_ces_vals.append(c_val)
                                elif "NPS" in ind_clean: prev_nps_vals.append(c_val)

                def avg(lst, is_c=False):
                    if not lst:
                        return "-"
                    return format_val(sum(lst) / len(lst), is_csat=is_c)

                if curr_csat_vals or curr_ces_vals or curr_nps_vals:
                    row_data = [
                        avg(curr_nps_vals, False), format_val(69, is_csat=False), avg(prev_nps_vals, False),
                        avg(curr_ces_vals, False), format_val(76, is_csat=False), avg(prev_ces_vals, False),
                        avg(curr_csat_vals, True), format_val(85, is_csat=True), avg(prev_csat_vals, True),
                        str(s), str(dept)
                    ]
                    annual_table_data.append([process_arabic_text(cell) for cell in row_data])

        # Style and append annual summary table element to PDF flow if data exists
        if len(annual_table_data) > 1:
            t_annual = Table(annual_table_data, colWidths=col_widths)
            t_annual.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#7b5ab8")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f4f6f9")),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            
            annual_block = KeepTogether([
                Paragraph(process_arabic_text("📊 جدول الملخص السنوي المجمع (بناءً على متوسط النصفين لكل سنة):"), section_title_style),
                t_annual
            ])
            elements.append(annual_block)
            elements.append(Spacer(1, 12))

    # --- Section B: Build PDF Individual Period Tables Section ---
    for p_name, rows in periods_data.items():
        matrix = build_service_matrix(rows)
        period_table_data = [headers]
        for item in matrix:
            row = [
                str(item["Current NPS"]), str(item["Target NPS"]), str(item["Prvious NPS"]),
                str(item["Current CES"]), str(item["Target CES"]), str(item["Prvious CES"]),
                str(item["Current CSAT"]), str(item["Target CSAT"]), str(item["Prvious CSAT"]),
                str(item["Service"]), str(item["Department"])
            ]
            period_table_data.append([process_arabic_text(cell) for cell in row])

        # Create and style table for each specific period
        t_period = Table(period_table_data, colWidths=col_widths)
        t_period.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5a3e8d")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f4f6f9")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        
        # Keep section title and its corresponding table grouped together
        period_block = KeepTogether([
            Paragraph(process_arabic_text(f" جدول بيانات الفترة: {p_name}"), section_title_style),
            t_period
        ])
        elements.append(period_block)
        elements.append(Spacer(1, 12))

    # Build the final PDF document layout using templates and elements list
    doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    
    return buffer.getvalue(), f"Report_{period.replace(' ', '_')}.pdf", "application/pdf"