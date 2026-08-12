import pandas as pd
import io
import os
import pymysql
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import load_workbook

def process_arabic_text(text):
    """
    Reshapes and reverses Arabic text using arabic_reshaper and bidi algorithm 
    so that Arabic characters render correctly from right to left in PDF outputs.
    """
    # Reshape Arabic characters to connect properly based on their positions
    reshaped_text = arabic_reshaper.reshape(str(text))
    # Apply BiDi algorithm to fix text direction flow for PDF display
    bidi_text = get_display(reshaped_text)
    return bidi_text

def get_previous_period(period_str):
    """
    Determines the previous year and quarter string accurately 
    (e.g., 'الربع الثاني 2026' -> 'الربع الأول 2026', or 'الربع الأول 2026' -> 'الربع الرابع 2025').
    """
    try:
        # Split the period string by separator to extract year and quarter text
        parts = period_str.split(" - ")
        year = int(parts[0].strip())
        q_text = parts[1].strip()
        
        # Define the list of standard quarters in order
        quarters = ["الربع الأول", "الربع الثاني", "الربع الثالث", "الربع الرابع"]
        if q_text in quarters:
            idx = quarters.index(q_text)
            # If it's not the first quarter, return the previous quarter of the same year
            if idx > 0:
                return year, quarters[idx - 1]
            else:
                # If it's the first quarter, return the fourth quarter of the previous year
                return year - 1, "الربع الرابع"
    except:
        pass
    return None, None

def fetch_specific_period_data(year_val, period_val):
    """
    Fetches raw indicator records and CSAT factor response data for a specific year and quarter 
    directly from the MySQL database.
    """
    try:
        # Establish connection to the MySQL database using specific credentials
        connection = pymysql.connect(
            host='localhost',
            port=3307,
            user='root',
            password='COOP@nllr2026',
            database='individuals_experience_db',
            charset='utf8mb4',
            cursorclass=pymysql.cursors.DictCursor
        )
        with connection.cursor() as cursor:
            # SQL query to fetch indicator responses and names for the given year and period
            query = """
                SELECT ind.IndicatorName, res.RatingValue AS CurrentValue
                FROM individualmeasurementrecords mr
                JOIN individualindicatorresponses res ON mr.RecordID = res.RecordID
                JOIN sharedindicators ind ON res.IndicatorID = ind.IndicatorID
                WHERE mr.Year = %s AND mr.Period = %s
            """
            cursor.execute(query, (year_val, period_val))
            rows = cursor.fetchall()
            
            # SQL query to fetch CSAT factor response records and names for the given year and period
            factor_query = """
                SELECT f.FactorName, fr.RatingValue
                FROM individualmeasurementrecords mr
                JOIN individualfactorresponses fr ON mr.RecordID = fr.RecordID
                JOIN sharedcsatfactors f ON fr.FactorID = f.FactorID
                WHERE mr.Year = %s AND mr.Period = %s
            """
            cursor.execute(factor_query, (year_val, period_val))
            f_rows = cursor.fetchall()
        # Close the database connection safely
        connection.close()
        return rows, f_rows
    except:
        return [], []

def calculate_metrics(rows, factors_rows):
    """
    Computes CES and NPS net scores (-100 to 100) and calculates 
    percentage scores (0 to 100%) for each of the 8 CSAT factors.
    """
    res_dict = {}
    
    # 1. Calculate CES (Customer Effort Score) net score ranging from -100 to 100
    ces_high = sum(1 for r in rows if "CES" in str(r.get('IndicatorName', '')).upper() and r.get('CurrentValue') is not None and float(r.get('CurrentValue', 0)) >= 4)
    ces_low = sum(1 for r in rows if "CES" in str(r.get('IndicatorName', '')).upper() and r.get('CurrentValue') is not None and float(r.get('CurrentValue', 0)) <= 2)
    ces_total = sum(1 for r in rows if "CES" in str(r.get('IndicatorName', '')).upper() and r.get('CurrentValue') is not None)
    res_dict["CES"] = ((ces_high - ces_low) / ces_total * 100) if ces_total > 0 else None
    
    # 2. Calculate NPS (Net Promoter Score) net score ranging from -100 to 100
    nps_p = sum(1 for r in rows if "NPS" in str(r.get('IndicatorName', '')).upper() and r.get('CurrentValue') is not None and float(r.get('CurrentValue', 0)) >= 9)
    nps_d = sum(1 for r in rows if "NPS" in str(r.get('IndicatorName', '')).upper() and r.get('CurrentValue') is not None and float(r.get('CurrentValue', 0)) <= 6)
    nps_total = sum(1 for r in rows if "NPS" in str(r.get('IndicatorName', '')).upper() and r.get('CurrentValue') is not None)
    res_dict["NPS"] = ((nps_p - nps_d) / nps_total * 100) if nps_total > 0 else None
    
    # 3. Group and calculate scores for each of the 8 CSAT factors as percentages (0 to 100%)
    factors_dict = {}
    for f in factors_rows:
        fname = f.get('FactorName', 'عامل')
        fval = f.get('RatingValue', 0)
        try:
            fval_num = float(fval)
        except:
            fval_num = 0.0
        if fname not in factors_dict:
            factors_dict[fname] = []
        factors_dict[fname].append(fval_num)
        
    for fname, vals in factors_dict.items():
        if vals:
            pos = sum(1 for v in vals if v >= 4)
            neg = sum(1 for v in vals if v <= 2)
            total = len(vals)
            csat_val = ((pos - neg) / total * 100) if total > 0 else 0.0
            if csat_val < 0: csat_val = 0.0
            res_dict[f"CSAT - {fname}"] = csat_val
        else:
            res_dict[f"CSAT - {fname}"] = None
            
    return res_dict

def build_summary_matrix(p_name, rows, factors_rows):
    """
    Builds a summary matrix matching current period metrics against 
    automatically fetched previous period metrics for CES, NPS, and CSAT factors.
    """
    summary_list = []
    
    # Step 1: Calculate metrics for the current period specified in the report
    curr_metrics = calculate_metrics(rows, factors_rows)
    
    # Step 2: Automatically identify and fetch data for the preceding quarter period
    prev_year, prev_period_val = get_previous_period(p_name)
    prev_metrics = {}
    if prev_year and prev_period_val:
        p_rows, p_f_rows = fetch_specific_period_data(prev_year, prev_period_val)
        if p_rows or p_f_rows:
            prev_metrics = calculate_metrics(p_rows, p_f_rows)
            
    # Append CES row data with current, target, and previous values
    c_ces = curr_metrics.get("CES")
    p_ces = prev_metrics.get("CES")
    summary_list.append({
        "Indicator": "CES",
        "Previous": f"{p_ces:.2f}" if p_ces is not None else "-",
        "Target": "76.00",
        "Current": f"{c_ces:.2f}" if c_ces is not None else "-"
    })
    
    # Append NPS row data with current, target, and previous values
    c_nps = curr_metrics.get("NPS")
    p_nps = prev_metrics.get("NPS")
    summary_list.append({
        "Indicator": "NPS",
        "Previous": f"{p_nps:.2f}" if p_nps is not None else "-",
        "Target": "69.00",
        "Current": f"{c_nps:.2f}" if c_nps is not None else "-"
    })
    
    # Append each CSAT factor row with its corresponding current, target, and previous values
    for key, c_val in curr_metrics.items():
        if key.startswith("CSAT -"):
            p_val = prev_metrics.get(key)
            summary_list.append({
                "Indicator": key,
                "Previous": f"{p_val:.2f}%" if p_val is not None else "-",
                "Target": "85.00%",
                "Current": f"{c_val:.2f}%" if c_val is not None else "-"
            })
            
    return summary_list

def generate_excel(period, periods_data, factors_data):
    """
    Generates an Excel workbook using pandas and openpyxl, creating individual 
    sheets for each period and applying custom gray-purple header styling.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        for p_name, rows in periods_data.items():
            f_rows = factors_data.get(p_name, [])
            matrix = build_summary_matrix(p_name, rows, f_rows)
            df_period = pd.DataFrame(matrix)
            # Reorder columns to match the desired format layout
            df_period = df_period[["Indicator", "Previous", "Target", "Current"]]
            safe_sheet_name = p_name.replace(" & ", "_")
            df_period.to_excel(writer, index=False, sheet_name=safe_sheet_name)

    buffer.seek(0)
    wb = load_workbook(buffer)
    
    # Define professional header fill colors, fonts, and center alignments for Excel
    header_fill = PatternFill(start_color="7A7F94", end_color="7A7F94", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    alignment_center = Alignment(horizontal="center", vertical="center")

    # Apply styles to header row cells across all workbook sheets
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = alignment_center

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    
    return final_buffer.getvalue(), f"Summary_Report_{period.replace(' ', '_')}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

def generate_pdf(period, periods_data, factors_data):
    """
    Generates a professionally styled PDF report document using ReportLab, 
    incorporating dynamic headers, footers, custom fonts, and data tables.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=(842, 595), rightMargin=40, leftMargin=40, topMargin=90, bottomMargin=30)
    elements = []
    
    # Register TrueType Arabic font if available on the system path
    font_path = "C:\\Windows\\Fonts\\arial.ttf"
    font_name = "Helvetica" 
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
        font_name = 'ArabicFont'
        
    styles = getSampleStyleSheet()
    
    def add_header_footer(canvas_obj, document):
        """Draws running header logos, titles, and bottom dividing lines on PDF pages."""
        canvas_obj.saveState()
        logo_path = "logo.png"
        if os.path.exists(logo_path):
            canvas_obj.drawImage(logo_path, 40, 525, width=120, height=45, preserveAspectRatio=True, mask='auto')
        
        canvas_obj.setFont(font_name, 12)
        canvas_obj.drawRightString(800, 545, process_arabic_text(f"ملخص مؤشرات تجربة العميل (الأفراد) - الفترة: {period}"))
        
        canvas_obj.setLineWidth(1)
        canvas_obj.setStrokeColor(colors.HexColor('#1e253c'))
        canvas_obj.line(40, 515, 800, 515)
        canvas_obj.restoreState()

    elements.append(Spacer(1, 10))

    # Define paragraph style for section headers
    section_title_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontName=font_name,
        fontSize=12,
        alignment=2,
        textColor=colors.HexColor('#7b5ab8'),
        spaceAfter=10,
        spaceBefore=10,
        keepWithNext=True
    )

    # Define table column headers and custom column widths
    headers = [
        process_arabic_text("Current Value"), 
        process_arabic_text("Target Value"), 
        process_arabic_text("Previous Value"), 
        process_arabic_text("Indicator / Factor Name")
    ]
    col_widths = [120, 120, 120, 382]

    # Iterate through periods and populate ReportLab tables
    for p_name, rows in periods_data.items():
        f_rows = factors_data.get(p_name, [])
        matrix = build_summary_matrix(p_name, rows, f_rows)
        
        period_table_data = [headers]
        for item in matrix:
            row = [
                str(item["Current"]), 
                str(item["Target"]), 
                str(item["Previous"]), 
                str(item["Indicator"])
            ]
            period_table_data.append([process_arabic_text(cell) for cell in row])

        # Build and style the period summary table flowable
        t_period = Table(period_table_data, colWidths=col_widths)
        t_period.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#5a3e8d")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor("#f4f6f9")),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        # Wrap section title and table together to avoid splitting across pages
        period_block = KeepTogether([
            Paragraph(process_arabic_text(f"ملخص المؤشرات والعوامل للفترة: {p_name}"), section_title_style),
            t_period
        ])
        elements.append(period_block)

    # Build the document layout using canvas headers/footers and element flowables
    doc.build(elements, onFirstPage=add_header_footer, onLaterPages=add_header_footer)
    
    return buffer.getvalue(), f"Summary_Report_{period.replace(' ', '_')}.pdf", "application/pdf"